import openpyxl
from openpyxl.drawing.image import Image
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from scipy.optimize import fsolve
from scipy import stats
from sklearn.metrics import r2_score, mean_squared_error
import argparse
import sys
from pathlib import Path
from datetime import datetime
import logging
import json
import io

def setup_logging(output_dir, verbose):
    """
    Set up logging configuration
    
    Parameters:
    -----------
    output_dir : Path
        Directory for log file output
    verbose : bool
        If True, set logging level to DEBUG
    
    Returns:
    --------
    logger : Logger
        Configured logger instance
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = output_dir / f'analysis_log_{timestamp}.txt'
    
    logger = logging.getLogger('ELISA_Analysis')
    logger.setLevel(logging.DEBUG)
    
    # Configure file handler
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    
    # Configure console handler
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG if verbose else logging.INFO)
    
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)
    
    logger.addHandler(fh)
    logger.addHandler(ch)
    
    return logger

def standardize_sample_names(sample_name):
    """
    Standardize sample names by removing replicate indicators
    
    Parameters:
    -----------
    sample_name : str or int or float
        Sample name to standardize
        
    Returns:
    --------
    str
        Standardized sample name
    """
    import re
    
    if isinstance(sample_name, (int, float)):
        return str(int(sample_name))
    
    sample_name = str(sample_name).strip()
    
    patterns = [
        r'^(.*?)[_-][0-9A-Za-z]+$',
        r'^(.*?)\([0-9A-Za-z]+\)$',
        r'^(.*?)\s+[0-9A-Za-z]+$',
        r'^(.*?)_Rep[0-9A-Za-z]+$',
        r'^(.*?)-Rep[0-9A-Za-z]+$',
    ]
    
    for pattern in patterns:
        match = re.match(pattern, sample_name)
        if match:
            return match.group(1)
    
    return sample_name

def validate_sample_names(sample_data):
    """Check consistency of sample names and return warnings"""
    warnings = []
    
    name_pairs = [(name, standardize_sample_names(name)) 
                 for name in sample_data['Sample']]
    
    similar_names = {}
    for orig_name, std_name in name_pairs:
        if std_name not in similar_names:
            similar_names[std_name] = []
        similar_names[std_name].append(orig_name)
    
    for std_name, orig_names in similar_names.items():
        if len(orig_names) > 1:
            warnings.append(
                f"Warning: Similar sample names detected: "
                f"{', '.join(orig_names)}"
                f" -> Will use '{std_name}' as unified name"
            )
    
    return warnings

def four_pl(x, A, B, C, D):
    """
    4-parameter logistic function
    
    Parameters:
    -----------
    x : array-like
        Input values
    A : float
        Maximum asymptote
    B : float
        Slope
    C : float
        EC50
    D : float
        Minimum asymptote
    """
    return D + (A - D) / (1.0 + (x / C) ** B)

def five_pl(x, A, B, C, D, E):
    """
    5-parameter logistic function
    
    Additional parameter E controls asymmetry
    """
    return D + (A - D) / (1.0 + (x / C) ** B) ** E

def linear_regression(x, y):
    """Perform linear regression and return the function and coefficients"""
    coef = np.polyfit(x, y, 1)
    return np.poly1d(coef), coef

def calculate_fit_metrics(y_true, y_pred, n_params):
    """
    Calculate various metrics for evaluating the fit quality
    
    Returns:
    --------
    dict
        Contains R2, Adjusted R2, AIC, BIC, and RMSE values
    """
    n = len(y_true)
    residuals = y_true - y_pred
    rss = np.sum(residuals**2)
    tss = np.sum((y_true - np.mean(y_true))**2)

    r2 = 1 - (rss/tss)
    adj_r2 = 1 - ((1-r2)*(n-1)/(n-n_params-1))
    aic = n * np.log(rss/n) + 2 * n_params
    bic = n * np.log(rss/n) + n_params * np.log(n)
    rmse = np.sqrt(np.mean(residuals**2))
    
    return {
        'R2': r2,
        'Adjusted_R2': adj_r2,
        'AIC': aic,
        'BIC': bic,
        'RMSE': rmse
    }

def get_initial_params(y_data, x_data):
    """Get initial parameters for 4PL and 5PL curve fitting"""
    A_init = np.max(y_data) * 1.05
    D_init = np.min(y_data) * 0.95
    B_init = 1.0
    mid_response = (A_init + D_init) / 2
    closest_idx = np.argmin(np.abs(y_data - mid_response))
    C_init = x_data[closest_idx]
    E_init = 1.0
    
    return {
        'A': A_init,
        'B': B_init,
        'C': C_init,
        'D': D_init,
        'E': E_init
    }

def fit_curve(x_data, y_data, method, init_params=None, verbose=False):
    """
    Fit curve using specified method
    
    Parameters:
    -----------
    method : str
        'linear', '4', or '5' for linear regression, 4PL, or 5PL respectively
    """
    try:
        if method == 'linear':
            fit_func, coef = linear_regression(x_data, y_data)
            y_fit = fit_func(x_data)
            n_params = 2
            formula = f"y = {coef[0]:.4f}x + {coef[1]:.4f}"
            return fit_func, coef, y_fit, n_params, formula
        elif method == '4':
            bounds = ([0, 0.5, 0, 0], [np.inf, 10, np.inf, np.inf])
            p0 = [init_params['A'], init_params['B'], init_params['C'], init_params['D']]
            popt, pcov = curve_fit(four_pl, x_data, y_data, p0=p0, bounds=bounds, maxfev=50000)
            y_fit = four_pl(x_data, *popt)
            n_params = 4
            formula = f"y = {popt[3]:.4f} + ({popt[0]:.4f} - {popt[3]:.4f}) / (1 + (x/{popt[2]:.4f})^{popt[1]:.4f})"
            return four_pl, popt, y_fit, n_params, formula
        else:
            bounds = ([0, 0.5, 0, 0, 0.5], [np.inf, 10, np.inf, np.inf, 5])
            p0 = [init_params['A'], init_params['B'], init_params['C'], init_params['D'], init_params['E']]
            popt, pcov = curve_fit(five_pl, x_data, y_data, p0=p0, bounds=bounds, maxfev=50000)
            y_fit = five_pl(x_data, *popt)
            n_params = 5
            formula = f"y = {popt[3]:.4f} + ({popt[0]:.4f} - {popt[3]:.4f}) / (1 + (x/{popt[2]:.4f})^{popt[1]:.4f})^{popt[4]:.4f}"
            return five_pl, popt, y_fit, n_params, formula
    except RuntimeError as e:
        raise RuntimeError(f"Fitting failed: {str(e)}")

def inverse_fit_func(fit_func, absorbance, method, popt):
    """Calculate concentration from absorbance using the fitted function"""
    if method == 'linear':
        a, b = popt
        return (absorbance - b) / a
    else:
        def func_to_solve(conc):
            if method == '4':
                return four_pl(conc, *popt) - absorbance
            elif method == '5':
                return five_pl(conc, *popt) - absorbance
            else:
                raise ValueError(f"Unsupported method: {method}")

        x0 = 0.1 if absorbance <= 0 else 1.0
        
        try:
            concentration = fsolve(func_to_solve, x0=x0, full_output=True, xtol=1.49012e-8, maxfev=1000)
            if concentration[2] == 1:
                return float(concentration[0][0])
            else:
                raise RuntimeError("Failed to converge")
        except:
            try:
                concentration = fsolve(func_to_solve, x0=100.0, full_output=True)
                if concentration[2] == 1:
                    return float(concentration[0][0])
                else:
                    return np.nan
            except:
                return np.nan

def read_input_file(file_path, sheet_name=None):
    """
    Read input file (Excel or CSV)
    
    Parameters:
    -----------
    file_path : str or Path
        Path to input file
    sheet_name : str, optional
        Sheet name for Excel files
        
    Returns:
    --------
    pd.DataFrame
        Data frame containing the input data
    str
        File type ('excel' or 'csv')
    """
    file_path = Path(file_path)
    file_type = 'excel' if file_path.suffix.lower() in ['.xlsx', '.xls'] else 'csv'
    
    try:
        if file_type == 'excel':
            df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
        else:
            df = pd.read_csv(file_path)
        return df, file_type
    except Exception as e:
        raise RuntimeError(f"Failed to read input file: {str(e)}")

def process_standard_data(df):
    """
    Process standard data and automatically determine number of replicates
    """
    empty_row_idx = df.index[df.iloc[:, 1:].isna().all(axis=1)].min()
    standard_data = df.iloc[0:empty_row_idx, 1:].apply(pd.to_numeric, errors='coerce')
    
    n_replicates = len(standard_data)
    
    concentrations = df.columns[1:].astype(float)
    mean_values = standard_data.mean(axis=0)
    sd_values = standard_data.std(axis=0, ddof=1) if n_replicates > 1 else pd.Series(0, index=mean_values.index)
    se_values = sd_values / np.sqrt(n_replicates) if n_replicates > 1 else pd.Series(0, index=mean_values.index)
    
    return concentrations, mean_values, sd_values, se_values, n_replicates

def process_sample_data(df, empty_row_idx, logger):
    """
    Process sample data with three levels of statistics:
    1. Individual measurements
    2. Replicate statistics (identical sample names)
    3. Group statistics (base names)
    """
    from scipy import stats as scipy_stats
    
    sample_data = df.iloc[empty_row_idx+1:, [0, 1]].dropna()
    sample_data.columns = ['Sample', 'Absorbance']
    sample_data['Sample'] = sample_data['Sample'].astype(str)
    
    logger.debug(f"Loaded sample data (first 5 rows):\n{sample_data.head().to_string()}")
    
    # Save original sample names
    sample_data['Original_Sample_Name'] = sample_data['Sample']
    
    # Create base name for grouping (for statistics)
    sample_data['Base_Sample_Name'] = sample_data['Sample'].apply(standardize_sample_names)
    
    logger.debug(f"Sample data after standardization (first 5 rows):\n{sample_data.head().to_string()}")
    
    # 1. Store results for each individual sample measurement
    individual_results = []
    for _, row in sample_data.iterrows():
        individual_results.append({
            'Sample': row['Sample'],  # Original sample name (with replicate identifier)
            'Base_Sample_Name': row['Base_Sample_Name'],  # Base name for grouping
            'Absorbance': row['Absorbance'],
            'Original_Names': [row['Sample']],  # Keep as list for consistency
            'Is_Group': False,  # Explicitly mark as individual sample
            'Is_Replicate': False  # Not a replicate statistic
        })
    
    # 2. Calculate replicate statistics (same exact sample name)
    replicate_stats = []
    for sample, group in sample_data.groupby('Sample'):
        # Only create replicate statistics if there are multiple measurements
        if len(group) > 1:
            mean_abs = group['Absorbance'].mean()
            sd_abs = group['Absorbance'].std(ddof=1)
            se_abs = sd_abs / np.sqrt(len(group))
            
            t_value = scipy_stats.t.ppf(0.975, len(group)-1)
            ci_95 = t_value * se_abs
            
            cv_percent = (sd_abs / mean_abs * 100) if mean_abs != 0 else 0
            
            base_name = group['Base_Sample_Name'].iloc[0]  # All entries have same base name
            
            replicate_stats.append({
                'Sample': sample,  # Original sample name with replicate number
                'Base_Sample_Name': base_name,
                'Absorbance_Mean': mean_abs,
                'Absorbance_SD': sd_abs,
                'Absorbance_SE': se_abs,
                'N': len(group),
                'CI_95': ci_95,
                'CV_Percent': cv_percent,
                'Original_Names': [sample] * len(group),
                'Is_Group': False,  # Not a base group statistic
                'Is_Replicate': True  # Mark as replicate statistic
            })
    
    # 3. Calculate group statistics (base sample name)
    group_stats = []
    for sample, group in sample_data.groupby('Base_Sample_Name'):
        original_names = group['Original_Sample_Name'].tolist()
        n = len(group)
        mean_abs = group['Absorbance'].mean()
        sd_abs = group['Absorbance'].std(ddof=1) if n > 1 else 0
        se_abs = sd_abs / np.sqrt(n) if n > 1 else 0
        
        if n > 1:
            t_value = scipy_stats.t.ppf(0.975, n-1)
            ci_95 = t_value * se_abs
        else:
            ci_95 = 0
        
        cv_percent = (sd_abs / mean_abs * 100) if n > 1 and mean_abs != 0 else 0
        
        group_stats.append({
            'Sample': sample,
            'Absorbance_Mean': mean_abs,
            'Absorbance_SD': sd_abs,
            'Absorbance_SE': se_abs,
            'N': n,
            'CI_95': ci_95,
            'CV_Percent': cv_percent,
            'Original_Names': original_names,
            'Is_Group': True,  # Mark as group statistic
            'Is_Replicate': False  # Not a replicate statistic
        })
    
    logger.debug(f"Number of individual samples: {len(individual_results)}")
    logger.debug(f"Number of replicate statistics: {len(replicate_stats)}")
    logger.debug(f"Number of group statistics: {len(group_stats)}")
    
    # Combine all results
    all_results = individual_results + replicate_stats + group_stats
    result_df = pd.DataFrame(all_results)
    
    logger.debug(f"Combined DataFrame shape: {result_df.shape}")
    logger.debug(f"Individual samples (first 3):\n{result_df[(~result_df['Is_Group']) & (~result_df['Is_Replicate'])].head(3).to_string() if not result_df.empty else 'Empty'}")
    logger.debug(f"Replicate statistics (first 3):\n{result_df[result_df['Is_Replicate']].head(3).to_string() if not result_df.empty else 'Empty'}")
    logger.debug(f"Group statistics (first 3):\n{result_df[result_df['Is_Group']].head(3).to_string() if not result_df.empty else 'Empty'}")
    
    return result_df

def validate_data_quality(stats, cv_threshold=15):
    """
    Check data quality and generate warnings
    
    Parameters:
    -----------
    stats : DataFrame
        Statistics for each sample
    cv_threshold : float
        CV threshold for warnings (percent)
    """
    warnings = []
    
    # Check group and replicate statistics for high CV
    if 'Is_Group' in stats.columns and 'Is_Replicate' in stats.columns:
        check_stats = stats[(stats['Is_Group'] == True) | (stats['Is_Replicate'] == True)].copy()
    elif 'Is_Group' in stats.columns:
        check_stats = stats[stats['Is_Group'] == True].copy()
    else:
        check_stats = stats.copy()
    
    high_cv_samples = check_stats[check_stats['CV_Percent'] > cv_threshold]
    if not high_cv_samples.empty:
        for _, row in high_cv_samples.iterrows():
            sample_type = "group" if row.get('Is_Group', False) else "replicate"
            warnings.append(
                f"Warning: {row['Sample']} ({sample_type}) has CV of {row['CV_Percent']:.1f}% (threshold: {cv_threshold}%)"
            )
    
    single_samples = check_stats[check_stats['N'] == 1]
    if not single_samples.empty:
        for _, row in single_samples.iterrows():
            warnings.append(
                f"Note: {row['Sample']} has only single measurement"
            )
    
    return warnings

def create_and_save_plot(concentrations, mean_values, sd_values, sample_stats, 
                        fit_func, method, popt, formula, output_png_path, return_bytes=False):
    """
    Create and save standard curve plot with sample data
    
    Parameters:
    -----------
    ...
    return_bytes : bool
        If True, returns the plot as bytes for Excel embedding
    
    Returns:
    --------
    bytes or None
        Plot as bytes if return_bytes is True, None otherwise
    """
    plt.figure(figsize=(10, 8))
    
    plt.errorbar(concentrations, mean_values, yerr=sd_values, 
                fmt='o', label="Standards (mean ± SD)", 
                capsize=5, color='blue', markersize=8)
    
    x_vals = np.logspace(np.log10(min(concentrations)), np.log10(max(concentrations)), 500)
    if method == 'linear':
        plt.plot(x_vals, fit_func(x_vals), '-', label="Fitted Curve (Linear)", 
                color='red', linewidth=2)
    else:
        plt.plot(x_vals, fit_func(x_vals, *popt), '-', 
                label=f"Fitted Curve ({method}PL)", 
                color='red', linewidth=2)
    
    # Plot group statistics only
    if 'Is_Group' in sample_stats.columns:
        plot_stats = sample_stats[sample_stats['Is_Group'] == True].copy()
    else:
        plot_stats = sample_stats.copy()
    
    for _, row in plot_stats.iterrows():
        if 'Concentration' in row and not pd.isna(row['Concentration']):
            plt.errorbar(row['Concentration'], row['Absorbance_Mean'],
                        yerr=row['Absorbance_SD'] if row['N'] > 1 else None,
                        fmt='s', label=f"{row['Sample']} (n={row['N']})",
                        capsize=5, markersize=8)
    
    plt.text(0.05, 0.95, formula, transform=plt.gca().transAxes, 
            fontsize=10, verticalalignment='top',
            bbox=dict(facecolor='white', alpha=0.8))
    
    plt.xscale('log')
    plt.xlabel('Concentration', fontsize=12)
    plt.ylabel('Absorbance', fontsize=12)
    plt.title('Standard Curve with Sample Data', fontsize=14)
    plt.grid(True, which="both", ls="-", alpha=0.2)
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=10)
    
    plt.tight_layout()
    
    # Save to file
    plt.savefig(output_png_path, dpi=300, bbox_inches='tight')
    
    # If requested, also return as bytes
    if return_bytes:
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        plt.close()
        return buffer.getvalue()
    
    plt.close()
    return None

def save_results(output_path, sample_stats, metrics, formula, concentrations, 
                mean_values, sd_values, se_values, n_std_replicates, warnings, 
                fit_func, method, popt, output_png_path, file_type='excel', logger=None):
    """
    Save analysis results to Excel/CSV file with three levels of results:
    1. Individual measurements
    2. Replicate statistics (same exact sample names)
    3. Group statistics (base names)
    """
    output_path = Path(output_path)
    
    if logger:
        logger.debug(f"Sample stats DataFrame shape: {sample_stats.shape}")
        logger.debug(f"Is_Group column exists: {'Is_Group' in sample_stats.columns}")
        logger.debug(f"Is_Replicate column exists: {'Is_Replicate' in sample_stats.columns}")
    
    if file_type == 'excel':
        workbook = openpyxl.Workbook()
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            if 'Is_Group' in sample_stats.columns and 'Is_Replicate' in sample_stats.columns:
                individual_results = sample_stats[(~sample_stats['Is_Group']) & (~sample_stats['Is_Replicate'])].copy()
            else:
                individual_results = pd.DataFrame()
                
            if not individual_results.empty:
                available_cols = individual_results.columns.tolist()
                desired_cols = ['Sample', 'Base_Sample_Name', 'Absorbance', 'Concentration']
                individual_cols = [col for col in desired_cols if col in available_cols]
                
                if logger:
                    logger.debug(f"Selected columns for individual results: {individual_cols}")
                
                individual_results = individual_results[individual_cols].copy()
                individual_results.to_excel(writer, sheet_name="Individual_Results", index=False)
                
                if logger:
                    logger.debug("Individual_Results sheet created successfully")
            
            if 'Is_Replicate' in sample_stats.columns:
                replicate_results = sample_stats[sample_stats['Is_Replicate'] == True].copy()
                
                if not replicate_results.empty:
                    drop_cols = ['Is_Group', 'Is_Replicate']
                    replicate_results = replicate_results.drop(columns=[col for col in drop_cols if col in replicate_results.columns])
                    replicate_results.to_excel(writer, sheet_name="Replicate_Statistics", index=False)
                    
                    if logger:
                        logger.debug("Replicate_Statistics sheet created successfully")
            
            if 'Is_Group' in sample_stats.columns:
                group_results = sample_stats[sample_stats['Is_Group'] == True].copy()
                
                if not group_results.empty and 'Original_Names' in group_results.columns:
                    group_results['Original_Sample_Names'] = group_results['Original_Names'].apply(lambda x: ', '.join(x))
                    drop_cols = ['Original_Names', 'Is_Group']
                    if 'Is_Replicate' in group_results.columns:
                        drop_cols.append('Is_Replicate')
                    group_results = group_results.drop(columns=[col for col in drop_cols if col in group_results.columns])
                    group_results.to_excel(writer, sheet_name="Group_Statistics", index=False)
                    
                    if logger:
                        logger.debug("Group_Statistics sheet created successfully")
            
            # Sample name mapping
            if 'Is_Group' in sample_stats.columns:
                group_results = sample_stats[sample_stats['Is_Group'] == True].copy()
                if not group_results.empty and 'Original_Names' in group_results.columns:
                    group_results['Original_Sample_Names'] = group_results['Original_Names'].apply(lambda x: ', '.join(x))
                    name_mapping_df = pd.DataFrame({
                        'Standardized_Name': group_results['Sample'],
                        'Original_Names': group_results['Original_Sample_Names']
                    })
                    name_mapping_df.to_excel(writer, sheet_name="Sample_Name_Mapping", index=False)
            
            # Standard data
            std_data = pd.DataFrame({
                'Concentration': concentrations,
                'Mean_Absorbance': mean_values,
                'SD': sd_values,
                'SE': se_values,
                'N': n_std_replicates
            })
            std_data.to_excel(writer, sheet_name="Standard_Data", index=False)
            
            # Fitting metrics
            metrics_df = pd.DataFrame({
                'Metric': ['R²', 'Adjusted R²', 'RMSE', 'AIC', 'BIC'],
                'Value': [metrics['R2'], metrics['Adjusted_R2'], 
                         metrics['RMSE'], metrics['AIC'], metrics['BIC']]
            })
            metrics_df.to_excel(writer, sheet_name="Fitting_Metrics", index=False)
            
            # Formula
            pd.DataFrame({'Formula': [formula]}).to_excel(writer, 
                sheet_name="Formula", index=False)
            
            # Warnings
            if warnings:
                pd.DataFrame({'Warnings': warnings}).to_excel(writer, 
                    sheet_name="Quality_Checks", index=False)
            
            # Add plot to a new sheet
            workbook = writer.book
            plot_sheet = workbook.create_sheet("Plot")
            
            # Generate plot
            plot_bytes = create_and_save_plot(
                concentrations, mean_values, sd_values, 
                sample_stats,
                fit_func, method, popt, formula, output_png_path, return_bytes=True
            )
            
            # Add image to Excel
            img = openpyxl.drawing.image.Image(io.BytesIO(plot_bytes))
            plot_sheet.add_image(img, 'A1')
            
            # Adjust column widths
            for ws in workbook.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
                    
            if logger:
                logger.debug(f"Created sheets: {workbook.sheetnames}")
    else:
        # CSV output with similar modifications
        output_str = io.StringIO()
        
        # 1. Individual sample results
        output_str.write("=== Individual Sample Results ===\n")
        if 'Is_Group' in sample_stats.columns and 'Is_Replicate' in sample_stats.columns:
            individual_results = sample_stats[(~sample_stats['Is_Group']) & (~sample_stats['Is_Replicate'])].copy()
        else:
            individual_results = pd.DataFrame()
            
        if not individual_results.empty:
            individual_cols = ['Sample', 'Base_Sample_Name', 'Absorbance', 'Concentration']
            individual_cols = [col for col in individual_cols if col in individual_results.columns]
            individual_results = individual_results[individual_cols].copy()
            individual_results.to_csv(output_str, index=False)
        
        # 2. Replicate statistics
        output_str.write("\n=== Replicate Statistics ===\n")
        if 'Is_Replicate' in sample_stats.columns:
            replicate_results = sample_stats[sample_stats['Is_Replicate'] == True].copy()
            
            if not replicate_results.empty:
                drop_cols = ['Is_Group', 'Is_Replicate']
                replicate_results = replicate_results.drop(columns=[col for col in drop_cols if col in replicate_results.columns])
                replicate_results.to_csv(output_str, index=False)
        
        # 3. Group statistics results
        output_str.write("\n=== Group Statistics ===\n")
        if 'Is_Group' in sample_stats.columns:
            group_results = sample_stats[sample_stats['Is_Group'] == True].copy()
            
            if not group_results.empty and 'Original_Names' in group_results.columns:
                group_results['Original_Sample_Names'] = group_results['Original_Names'].apply(lambda x: ', '.join(x))
                drop_cols = ['Original_Names', 'Is_Group']
                if 'Is_Replicate' in group_results.columns:
                    drop_cols.append('Is_Replicate')
                group_results = group_results.drop(columns=[col for col in drop_cols if col in group_results.columns])
                group_results.to_csv(output_str, index=False)
        
        # Write standard data
        output_str.write("\n=== Standard Data ===\n")
        std_data = pd.DataFrame({
            'Concentration': concentrations,
            'Mean_Absorbance': mean_values,
            'SD': sd_values,
            'SE': se_values,
            'N': n_std_replicates
        })
        std_data.to_csv(output_str, index=False)
        
        # Write fitting metrics
        output_str.write("\n=== Fitting Metrics ===\n")
        metrics_str = (
            f"R²: {metrics['R2']:.6f}\n"
            f"Adjusted R²: {metrics['Adjusted_R2']:.6f}\n"
            f"RMSE: {metrics['RMSE']:.6e}\n"
            f"AIC: {metrics['AIC']:.6f}\n"
            f"BIC: {metrics['BIC']:.6f}\n"
        )
        output_str.write(metrics_str)
        
        # Write formula
        output_str.write("\n=== Formula ===\n")
        output_str.write(formula + "\n")
        
        # Write warnings
        if warnings:
            output_str.write("\n=== Quality Checks ===\n")
            output_str.write("\n".join(warnings))
        
        # Save to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(output_str.getvalue())
        
        output_str.close()

def process_data_and_calculate_conc(file_path, sheet_name, output_path, method, logger, verbose=False):
    """Main function to process ELISA data and calculate concentrations"""
    try:
        # Prepare text file for storing calculation process
        output_dir = Path(output_path).parent
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        calculation_log_path = output_dir / f'calculation_process_{timestamp}.txt'
        
        # Function to record calculation process
        def log_to_file(message):
            with open(calculation_log_path, 'a', encoding='utf-8') as f:
                f.write(f"{message}\n")
        
        # Record header information
        log_to_file("=" * 80)
        log_to_file(f"ELISA Analysis Calculation Process")
        log_to_file(f"Analysis started at: {timestamp}")
        log_to_file(f"Input file: {file_path}")
        log_to_file(f"Fitting method: {method}")
        log_to_file("=" * 80 + "\n")

        logger.info(f"Starting analysis: {file_path}")
        log_to_file("1. Reading input data...")
        
        # Read data
        df, file_type = read_input_file(file_path, sheet_name)
        
        # Process standard data
        log_to_file("\n2. Processing standard data:")
        concentrations, mean_values, sd_values, se_values, n_std_replicates = process_standard_data(df)
        
        # Record standard data details
        log_to_file("\nStandard data summary:")
        for conc, mean, sd, se in zip(concentrations, mean_values, sd_values, se_values):
            log_to_file(f"Concentration {conc:8.3f}: Mean = {mean:.4f}, SD = {sd:.4f}, SE = {se:.4f}")
        
        # Fitting process
        log_to_file("\n3. Curve fitting:")
        if method == 'auto':
            log_to_file("\nPerforming automatic method selection...")
            method, metrics, y_fit, popt, formula, fit_func = determine_best_fit(
                concentrations, mean_values, logger, verbose)
            log_to_file(f"Selected method: {method}")
        else:
            log_to_file(f"\nUsing specified method: {method}")
            init_params = get_initial_params(mean_values, concentrations) if method in ['4', '5'] else None
            if init_params:
                log_to_file("\nInitial parameters:")
                for param, value in init_params.items():
                    log_to_file(f"  {param}: {value:.6f}")
            
            fit_func, popt, y_fit, n_params, formula = fit_curve(
                concentrations, mean_values, method, init_params, verbose)
            metrics = calculate_fit_metrics(mean_values, y_fit, n_params)
            
        # Record fitting results
        log_to_file("\n4. Fitting results:")
        log_to_file(f"Method: {method}")
        log_to_file(f"Formula: {formula}")
        log_to_file("\nMetrics:")
        log_to_file(f"  R² = {metrics['R2']:.6f}")
        log_to_file(f"  Adjusted R² = {metrics['Adjusted_R2']:.6f}")
        log_to_file(f"  RMSE = {metrics['RMSE']:.6e}")
        log_to_file(f"  AIC = {metrics['AIC']:.6f}")
        log_to_file(f"  BIC = {metrics['BIC']:.6f}")

        # Process sample data - now with three levels of statistics
        empty_row_idx = df.index[df.iloc[:, 1:].isna().all(axis=1)].min()
        log_to_file("\n5. Processing sample data:")
        sample_stats = process_sample_data(df, empty_row_idx, logger)
        
        logger.debug(f"Sample_stats columns: {sample_stats.columns.tolist()}")
        
        # Calculate concentrations for each level
        log_to_file("\n6. Calculating concentrations:")
        
        # 1. Calculate for individual samples
        ind_mask = (~sample_stats['Is_Group']) & (~sample_stats['Is_Replicate'])
        if 'Absorbance' in sample_stats.columns and any(ind_mask):
            logger.debug("Calculating concentrations for individual samples")
            sample_stats.loc[ind_mask, 'Concentration'] = \
                sample_stats.loc[ind_mask, 'Absorbance'].apply(
                lambda x: inverse_fit_func(fit_func, x, method, popt)
            )
            
            logger.debug(f"Individual samples after concentration calculation (first 3):\n{sample_stats[ind_mask].head(3).to_string()}")
        
        # 2. Calculate for replicate statistics
        rep_mask = sample_stats['Is_Replicate'] == True
        if 'Absorbance_Mean' in sample_stats.columns and any(rep_mask):
            logger.debug("Calculating concentrations for replicate statistics")
            sample_stats.loc[rep_mask, 'Concentration'] = \
                sample_stats.loc[rep_mask, 'Absorbance_Mean'].apply(
                lambda x: inverse_fit_func(fit_func, x, method, popt)
            )
            
            # Calculate concentration statistics for replicates
            for stat in ['SD', 'SE']:
                stat_col = f'Absorbance_{stat}'
                if stat_col in sample_stats.columns:
                    sample_stats.loc[rep_mask, f'Concentration_{stat}'] = sample_stats[rep_mask].apply(
                        lambda row: (
                            inverse_fit_func(fit_func, row['Absorbance_Mean'] + row[stat_col], method, popt) -
                            inverse_fit_func(fit_func, row['Absorbance_Mean'] - row[stat_col], method, popt)
                        ) / 2 if not pd.isna(row[stat_col]) else np.nan,
                        axis=1
                    )
            
            if 'Absorbance_SE' in sample_stats.columns and 'CI_95' in sample_stats.columns:
                sample_stats.loc[rep_mask, 'Concentration_CI_95'] = sample_stats[rep_mask].apply(
                    lambda row: (
                        inverse_fit_func(fit_func, row['Absorbance_Mean'] + row['CI_95'], method, popt) -
                        inverse_fit_func(fit_func, row['Absorbance_Mean'] - row['CI_95'], method, popt)
                    ) / 2 if row['N'] > 1 and not pd.isna(row['CI_95']) else np.nan,
                    axis=1
                )
            
            logger.debug(f"Replicate statistics after concentration calculation (first 3):\n{sample_stats[rep_mask].head(3).to_string()}")
        
        # 3. Calculate for group statistics
        group_mask = sample_stats['Is_Group'] == True
        if 'Absorbance_Mean' in sample_stats.columns and any(group_mask):
            logger.debug("Calculating concentrations for group statistics")
            sample_stats.loc[group_mask, 'Concentration'] = \
                sample_stats.loc[group_mask, 'Absorbance_Mean'].apply(
                lambda x: inverse_fit_func(fit_func, x, method, popt)
            )
            
            # Calculate concentration statistics for groups
            for stat in ['SD', 'SE']:
                stat_col = f'Absorbance_{stat}'
                if stat_col in sample_stats.columns:
                    sample_stats.loc[group_mask, f'Concentration_{stat}'] = sample_stats[group_mask].apply(
                        lambda row: (
                            inverse_fit_func(fit_func, row['Absorbance_Mean'] + row[stat_col], method, popt) -
                            inverse_fit_func(fit_func, row['Absorbance_Mean'] - row[stat_col], method, popt)
                        ) / 2 if not pd.isna(row[stat_col]) else np.nan,
                        axis=1
                    )
            
            if 'Absorbance_SE' in sample_stats.columns and 'CI_95' in sample_stats.columns:
                sample_stats.loc[group_mask, 'Concentration_CI_95'] = sample_stats[group_mask].apply(
                    lambda row: (
                        inverse_fit_func(fit_func, row['Absorbance_Mean'] + row['CI_95'], method, popt) -
                        inverse_fit_func(fit_func, row['Absorbance_Mean'] - row['CI_95'], method, popt)
                    ) / 2 if row['N'] > 1 and not pd.isna(row['CI_95']) else np.nan,
                    axis=1
                )
            
            logger.debug(f"Group statistics after concentration calculation (first 3):\n{sample_stats[group_mask].head(3).to_string()}")
        
        # Record sample results at all three levels
        log_to_file("\nSample results:")
        
        # 1. Record individual sample results
        individual_results = sample_stats[(~sample_stats['Is_Group']) & (~sample_stats['Is_Replicate'])].copy()
        if not individual_results.empty:
            log_to_file("\nIndividual sample measurements:")
            for _, row in individual_results.head(5).iterrows():  # Just show first 5 to avoid very long logs
                log_to_file(f"Sample: {row['Sample']}")
                log_to_file(f"  Absorbance: {row['Absorbance']:.4f}")
                if 'Concentration' in row and not pd.isna(row['Concentration']):
                    log_to_file(f"  Calculated concentration: {row['Concentration']:.4f}")
                else:
                    log_to_file(f"  Calculated concentration: N/A")
            if len(individual_results) > 5:
                log_to_file(f"  ... plus {len(individual_results)-5} more individual measurements")
        
        # 2. Record replicate statistics results
        replicate_results = sample_stats[sample_stats['Is_Replicate'] == True].copy()
        if not replicate_results.empty:
            log_to_file("\nReplicate statistics:")
            for _, row in replicate_results.iterrows():
                log_to_file(f"\nSample replicate: {row['Sample']}")
                log_to_file(f"  Absorbance: {row['Absorbance_Mean']:.4f} ± {row['Absorbance_SD']:.4f}")
                log_to_file(f"  Number of measurements: {row['N']}")
                if 'Concentration' in row and not pd.isna(row['Concentration']):
                    log_to_file(f"  Calculated concentration: {row['Concentration']:.4f}")
                    if 'Concentration_SD' in row and not pd.isna(row['Concentration_SD']):
                        log_to_file(f"  Concentration SD: {row['Concentration_SD']:.4f}")
                else:
                    log_to_file(f"  Calculated concentration: N/A")
        
        # 3. Record group statistics results
        group_results = sample_stats[sample_stats['Is_Group'] == True].copy()
        if not group_results.empty:
            log_to_file("\nGroup statistics:")
            for _, row in group_results.iterrows():
                log_to_file(f"\nSample group: {row['Sample']}")
                log_to_file(f"  Absorbance: {row['Absorbance_Mean']:.4f} ± {row['Absorbance_SD']:.4f}")
                log_to_file(f"  Number of measurements: {row['N']}")
                if 'Concentration' in row and not pd.isna(row['Concentration']):
                    log_to_file(f"  Calculated concentration: {row['Concentration']:.4f}")
                    if 'Concentration_SD' in row and not pd.isna(row['Concentration_SD']):
                        log_to_file(f"  Concentration SD: {row['Concentration_SD']:.4f}")
                else:
                    log_to_file(f"  Calculated concentration: N/A")
            
        # Data quality check
        warnings = validate_data_quality(sample_stats)
        if warnings:
            log_to_file("\n7. Quality warnings:")
            for warning in warnings:
                log_to_file(f"  {warning}")
        
        # Create plot
        plot_png_path = Path(output_path).with_suffix('.png')
        log_to_file("\n8. Creating plots and saving results...")
        create_and_save_plot(
            concentrations, mean_values, sd_values, sample_stats,
            fit_func, method, popt, formula, plot_png_path
        )
        
        # Save results
        save_results(
            output_path, sample_stats, metrics, formula,
            concentrations, mean_values, sd_values, se_values,
            n_std_replicates, warnings, fit_func, method, popt,
            plot_png_path, file_type=file_type, logger=logger
        )
        
        # Completion message
        log_to_file("\n" + "=" * 80)
        log_to_file("Analysis completed successfully")
        log_to_file(f"Results saved to: {output_path}")
        log_to_file(f"Plot saved to: {plot_png_path}")
        log_to_file(f"Calculation process log saved to: {calculation_log_path}")
        log_to_file("=" * 80)
        
        logger.info("\nAnalysis completed")
        logger.info(f"Results file: {output_path}")
        logger.info(f"Plot image: {plot_png_path}")
        logger.info(f"Calculation log: {calculation_log_path}")
        
        return len(sample_stats)
    
    except Exception as e:
        error_msg = f"Error occurred during processing: {str(e)}"
        logger.error(error_msg)
        log_to_file(f"\nERROR: {error_msg}")
        raise

def determine_best_fit(concentrations, mean_values, logger, verbose=False):
    """
    Determine the best fitting method based on AIC and RMSE
    """
    methods = ['linear', '4', '5']
    best_method = None
    best_metrics = None
    best_y_fit = None
    best_popt = None
    best_formula = ""
    best_fit_func = None
    
    for method in methods:
        try:
            init_params = get_initial_params(mean_values, concentrations) if method in ['4', '5'] else None
            fit_func, popt, y_fit, n_params, formula = fit_curve(concentrations, mean_values, method, init_params, verbose)
            metrics = calculate_fit_metrics(mean_values, y_fit, n_params)

            if best_metrics is None or metrics['AIC'] < best_metrics['AIC']:
                best_method = method
                best_metrics = metrics
                best_y_fit = y_fit
                best_popt = popt
                best_formula = formula
                best_fit_func = fit_func
            elif metrics['AIC'] == best_metrics['AIC'] and metrics['RMSE'] < best_metrics['RMSE']:
                best_method = method
                best_metrics = metrics
                best_y_fit = y_fit
                best_popt = popt
                best_formula = formula
                best_fit_func = fit_func

        except RuntimeError:
            logger.warning(f"Fitting failed for {method} method")
    
    return best_method, best_metrics, best_y_fit, best_popt, best_formula, best_fit_func

def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(description='ELISA Analysis Tool')
    parser.add_argument('--input', '-i', required=True, help='Input Excel/CSV file')
    parser.add_argument('--sheet', '-s', help='Sheet name (for Excel files)')
    parser.add_argument('--method', '-m', choices=['4', '5', 'linear', 'auto'], 
                       default='4', help='Fitting method (4: 4PL, 5: 5PL, linear, auto)')
    parser.add_argument('--output', '-o', help='Output file path (default: results_[input_name].[xlsx/csv])')
    parser.add_argument('--output-format', '-f', choices=['excel', 'csv'], 
                       help='Output format (default: same as input)')
    parser.add_argument('--verbose', '-v', action='store_true', help='Display detailed output')
    args = parser.parse_args()

    input_path = Path(args.input)
    
    # Determine input and output formats
    input_format = 'excel' if input_path.suffix.lower() in ['.xlsx', '.xls'] else 'csv'
    output_format = args.output_format or input_format
    
    # Set default output path if not specified
    if not args.output:
        output_suffix = '.xlsx' if output_format == 'excel' else '.csv'
        output_path = input_path.parent / f'results_{input_path.stem}{output_suffix}'
    else:
        output_path = Path(args.output)
    
    # Always enable verbose mode for debugging
    args.verbose = True
    logger = setup_logging(input_path.parent, args.verbose)
    logger.info("DEBUG MODE ENABLED - Detailed log information will be shown")

    try:
        num_samples = process_data_and_calculate_conc(
            args.input, args.sheet, output_path, args.method, logger, args.verbose
        )
        logger.info(f"Processing completed: analyzed {num_samples} samples")
        logger.info("\nNOTE: Please check all worksheets in the Excel file. This version creates separate sheets for:")
        logger.info("- Individual_Results: Contains each individual measurement")
        logger.info("- Replicate_Statistics: Contains statistics for identical sample names (duplicates)")
        logger.info("- Group_Statistics: Contains statistics for each base sample group")
        return 0
    except Exception as e:
        logger.error(f"Error occurred: {str(e)}")
        return 1

if __name__ == '__main__':
    exit(main())
